'''
Created on Nov 26, 2017
@author Bryan Howell, Ph.D.

Description of script.
This script reads a raw xlxs export from Eduphoria and converts the data to 
a csv file for Ndatis
'''

''' ========== Python Modules ========== '''

# base open-source python modules
import numpy as np
import openpyxl
import glob
import time

# Bryan Howell's python modules
import csvwrap as csvw
import procasmt as pasmt


''' ==================== User Input ==================== '''
# NOTE: search for "START HERE" to see where script begins
# (code between here and "START HERE" is for fxn definitions)

# teacher info. (required for correct file names)
teacher = 'lewis_a_howell'
course = 'algebra2'
period = 'period8'
asmtdate = '11022017'

# use roster (yes=True and no=False)
deid = False

# pertinent directories and files
datadir = '/home/bryan/Desktop/edusoft_datatools/data_conversion/test6'
rostdir = datadir+'/master_roster'
rostfile = 'Book1.xlsx'

''' ==================== Functions ==================== '''


def ckint(s):
    '''ckint determines if a string is an integer
    Parameters:
    s, input string
    Returns:
    True (if integer) or False (it not)
    '''
    try:
        int(s)
        return True
    except ValueError:
        return False


def ckfloat(s):
    '''ckfloat determines if a string is a floating point number
    Parameters:
    s, input string
    Returns:
    True (if integer) or False (it not)
    '''
    try:
        float(s)
        return True
    except ValueError:
        return False


def ele_in_list(e, reflist):
    '''ele_in_list checks to see if element is in a list
    Parameters:
    e, element
    Return:
    ie, index in list where e appears
    '''
    try:
        return reflist.index(e)
    except:
        return None


def transplist(l):
    '''transplist executes transpose operator on a list
    Parameters:
    l, list before transposition
    Returns:
    lT, list after transposition
    '''

    return list(map(list, zip(*l)))


def readroster(rfile):
    '''readroster read data from a roster

    Parameters:
    rfile, roster file
    Returns:
    names, names of all students to reference
    ids, ids of the respective students
    '''
    # open file
    book = openpyxl.load_workbook(rfile)
    sheet = book.active  # sheet 1 (default)

    # determine if file is of the right size
    row_range = np.array([sheet.min_row, sheet.max_row])
    col_range = np.array([sheet.min_column, sheet.max_column])
    if(col_range[1]-col_range[0]+1 < 2):
        print('roster is missing columns.')
    numstud = row_range[1]-row_range[0]+1

    # preallocate data
    names = ['']*numstud
    ids = ['']*numstud

    # extract info. from roster
    # 1-based counting for xlxs files
    for ii in range(row_range[0], row_range[1]+1):
        name_i = sheet.cell(row=ii, column=1).value
        id_i = sheet.cell(row=ii, column=2).value
        if(name_i is not None):
            names[ii-1] = name_i.strip()
        if(id_i is not None):
            ids[ii-1] = str(id_i).strip()

    return names, ids


def sizedata(sheet):
    '''sizedata determines the first and last cells of the data
    Parameters:
    sheet, Excel worksheet
    Returns:
    cell_o, (row,column) of first cell
    cell_f, (row,column) of last cell
    '''

    cell_o = np.array([sheet.min_row, sheet.min_column])
    cell_f = np.array([sheet.max_row, sheet.max_column])

    return cell_o, cell_f


def find_names(sheet, cell_o, cell_f):
    '''find_names finds and extracts student names from data
    Parameters:
    sheet, Excel worksheet 
    cell_o, (row,column) of first cell
    cell_f, (row,column) of last cell       
    Returns:
    loc_snames, (row,column) of student names
    '''

    # find start of student names
    loc_snames = np.ones((2))
    snames_found = False

    for ii in range(cell_o[0], cell_f[1]+1):
        for jj in range(cell_o[0], cell_f[1]+1):
            cell_ij = sheet.cell(row=ii, column=jj).value
            if(type(cell_ij) is str):
                if(cell_ij.strip() == 'Student Name'):
                    loc_snames = np.array([ii, jj])
                    snames_found = True
                    break

    # stop here if student names were not found
    if(snames_found == False):
        print('Unable to locate student names')
        exit()

    return loc_snames


def find_respscore(sheet, cell_sname, cell_f):
    '''find_resp the location where students'responses and scores begin
    Parameters:
    sheet, Excel worksheet 
    cell_sname, (row,colmn) where header of student names are found  
    cell_f, (row,column) of last cell        
    Returns:
    loc_resp, (row,column) of beginning of responses
    loc_score, (row,column) of beginning of raw scores    
    '''

    # find (row,column) of responses, and raw and percent scores
    loc_resp = np.ones((2))
    loc_score = np.ones((2))
    resp_found = False
    score_found = False

    for jj in range(cell_sname[1], cell_f[1]+1):
        cell_ij = sheet.cell(row=cell_sname[0], column=jj).value
        if(type(cell_ij) is str):
            if(cell_ij.strip() == 'All Responses'):
                loc_resp = np.array([cell_sname[0], jj])
                resp_found = True
            if(cell_ij.strip() == 'Raw Score'):
                loc_score = np.array([cell_sname[0], jj])
                score_found = True
                break

    # stop here if responses or scores were not found
    if(not resp_found or not score_found):
        if(not resp_found):
            print('Unable to locate responses')
        if(not resp_found):
            print('Unable to raw scores')
        exit()

    return loc_resp, loc_score


def extrdata(sheet):
    '''extrdata finds the (row,col) location of pertinent data
    Parameters:
    sheet, Excel worksheet
    Returns:
    d, dictionary containing the following fields:
    *names, student names
    *responses, responses to questions in asmt
    *teks, TEKS associated w/ each question
    *raw scores, point value earned
    *percent scores, points earned as percentage of max. value 
    '''

    d = dict()  # preallocate dictionary

    # find cells that pinpoint where data is stored
    firstcell, lastcell = sizedata(sheet)  # first and last cells for all data
    loc_snames = find_names(sheet, firstcell, lastcell)  # student names
    loc_resp, loc_scores = find_respscore(
        sheet, loc_snames, lastcell)  # responses/scores

    # find first non-empty row
    rowstart = 0
    # FIXED MISTAKE (BH, 01/14/2018)
    for ii in range(loc_snames[0]+1, lastcell[0]+1):
        if(type(sheet.cell(row=ii, column=loc_snames[1]).value) is str):
            rowstart = ii
            break

    # rows and columns of interest
    # NOTE: range = [a,b) => the right value is non-inclusive
    datarow = range(rowstart, lastcell[0]+1)
    datacol = range(loc_resp[1], loc_scores[1])
    nstud = len(datarow)
    nques = len(datacol)

    # extract student names and scores
    sname = ['']*nstud
    score = ['']*nstud
    perc = ['']*nstud
    for ii in range(nstud):
        cell_stud = sheet.cell(row=datarow[ii], column=loc_snames[1]).value
        cell_score = sheet.cell(row=datarow[ii], column=loc_scores[1]).value
        cell_perc = sheet.cell(row=datarow[ii], column=loc_scores[1]+1).value
        if(cell_stud is not None):
            sname[ii] = cell_stud
        if(cell_score is not None):
            score[ii] = int(cell_score)
        if(cell_perc is not None):
            if(type(cell_perc) is str):
                perc[ii] = float(cell_perc.split('%')[0])/100
            else:
                perc[ii] = cell_perc

    # extract TEKS
    # first, find the first row of the TEKS
    tekrow = 0
    for ii in range(loc_snames[0]+1, lastcell[0]):
        cell_ij = sheet.cell(row=ii, column=loc_resp[1]).value
        if(type(cell_ij) is str):
            if('.' in cell_ij):
                tekrow = ii
                break

    # extract TEKS
    teks = ['']*nques
    for jj in range(nques):
        cell_tek = sheet.cell(row=tekrow, column=datacol[jj]).value
        if(type(cell_tek) is str):
            tmp = cell_tek.split('.')
            teks[jj] = tmp[0]+'.'+tmp[1].split('[')[0].strip()

    # extract responses
    data = [['']*nques for ii in range(nstud)]
    for ii in range(nstud):
        for jj in range(nques):
            cell_ij = sheet.cell(row=datarow[ii], column=datacol[jj]).value
            if(cell_ij is not None):
                data[ii][jj] = str(cell_ij)

    d['names'] = sname
    d['responses'] = data
    d['teks'] = teks
    d['raw scores'] = score
    d['percent scores'] = perc

    return d


def buildhead(X, teks, rscore, pscore):
    '''buildhead constructs the header for NDatis file formats

    Parameters:
    X, student responses
    teks, TEKS associated w/ each question
    rscore, raw scores (value x (# MC + # GR) + credit from FR)
    pscore, percent scores (max. points earned / max. points possible)

    Returns:
    hdr, the header for the Ndatis file
    * TEK, the take for each question
    * value, the value assigned to each question
    * answer, the answer
    * type, the type of question (MC, FR, or GR)
        MC = multiple choice
        FR = free response
        GR = gridded response    
    '''

    # size of data
    m = len(X)
    n = len(X[0])
    # preallocate
    qval = ['']*n
    qans = ['']*n
    qtype = ['']*n

    # UPDATE U.1 (BH, 01/14/2018) - improved categorization of questions
    # first, look for multiple choice questions
    for jj in range(n):

        lastplus = -1  # last row w/ + sign (0-based index)
        numplus = 0  # number of + signs
        numint = 0  # number of integers
        allsame = True  # are all answers the same?
        for ii in range(m):
            if(len(X[ii][jj]) > 0):
                if(X[ii][jj][0] == '+'):  # check if cell has + sign
                    numplus += 1
                    lastplus = ii
                if(ckint(X[ii][jj]) == True):  # check if cell has an integer
                    numint += 1
                if(not X[ii][jj] == X[0][jj]):  # check if all cell contents are the same
                    allsame = False

        if(numplus == m or numint == m):  # FR (most likely) or all questions correct

            xij_noplus = X[lastplus][jj].strip('+')
            if(xij_noplus.isalpha()):
                qtype[jj] = 'MC'
                qans[jj] = xij_noplus
            elif(numint == m and not allsame):
                qtype[jj] = 'FR'
            else:  # if not, FR or GR
                qtype[jj] = 'GR'

        elif(lastplus > -1):  # correct answer found (MC or GR)

            xij_noplus = X[lastplus][jj].strip('+')
            if(xij_noplus.isalpha()):
                qtype[jj] = 'MC'
                qans[jj] = xij_noplus
            else:
                qtype[jj] = 'GR'
                qans[jj] = xij_noplus

        else:  # no answer found (MC or GR) -- FR always has a +

            qans[jj] = 'N/A'
            if(X[m-1][jj].isalpha()):
                qtype[jj] = 'MC'
            else:
                qtype[jj] = 'GR'
    # END OF UPDATE U.1

    # grade asmt
    Xg = np.zeros((m, n))
    for jj in range(n):
        if(qtype[jj] == 'MC' or qtype[jj] == 'GR'):
            for ii in range(m):
                xij = str(X[ii][jj])
                if(len(xij) > 0):
                    if(xij[0] == '+'):
                        Xg[ii, jj] = 1
        else:
            for ii in range(m):
                xij = str(X[ii][jj])
                if(len(xij) > 0):
                    Xg[ii, jj] = int(X[ii][jj])

    # determine values for questions
    qindx = pasmt.parse_quest(qtype)
    bin_indx = np.sort(np.concatenate(
        (qindx.get('mc'), qindx.get('gr'))))  # binary
    par_indx = qindx.get('fr')  # partial credit
    numbin = len(bin_indx)
    numpar = len(par_indx)

    # find first row w/ non-zero row
    rowtry = 0
    for ii in range(m):
        if(not pscore[ii] == 0):  # FIXED MISTAKE (BH, 01/14/2018)
            rowtry = ii
            break
    maxscore = int(round(rscore[rowtry]/pscore[rowtry]))
    ptval = maxscore/n

#     if(numbin>0 and numbin<=m): # if enough data present...
#         # use pseudoinverse to estimate accurately the values
#         # allows for different values/weights
#         A=Xg[:,bin_indx]
#         y=np.asarray(rscore)-np.sum(Xg[:,par_indx],axis=1)
#         Astar=np.linalg.pinv(A)
#         v_bin=np.matmul(Astar,y)
#         for ii in range(numbin):
#             qval[bin_indx[ii]]=round(v_bin[ii],2)
#         if(numpar>0):
#             v_par=(maxscore-np.sum(v_bin))/numpar
#             for ii in range(numpar):
#                 qval[par_indx[ii]]=round(v_par,2)
#     else:
#         print()

    # uniform weighting is assumed
    for ii in range(n):
        qval[ii] = ptval

    # determine max. credit for FR
    if(numpar > 0):

        # convert list of values to an array
        ptval = np.asarray(qval)
        pt_bin = ptval[bin_indx]  # points for MC/GR (i.e., binary questions)
        pt_par = ptval[par_indx]  # points for FR (i.e., partial credit)
        T = np.sum(ptval)  # total points that can be earned
        if(numbin > 0):
            Tbin = np.sum(np.tile(pt_bin, (m, 1))*Xg[:, bin_indx], axis=1)
        else:
            Tbin = 0
        # partial credit earned for all students
        Tpar = T*np.asarray(pscore)-Tbin
        for ii in range(m):
            Tpar[ii] = round(Tpar[ii])
        nzindx_Tpar = np.where(Tpar > 0)[0]  # where partial credit is non-zero

        # find enough rows to search for unknown max. credit
        if(len(nzindx_Tpar) < numpar):
            print('Not enough data.')
            exit()
        else:
            rws = nzindx_Tpar[0:numpar]

        # use Newton-Raphson method to find roots of nonlinear eqn.
        xmax = np.amax(Xg[:, par_indx], axis=0)
        tmp = Xg[rws, :]
        Xg_sub = tmp[:, par_indx]
        ptpar_rp = np.tile(pt_par, (numpar, 1))
        R = np.sum(ptpar_rp*Xg_sub/np.tile(xmax,
                                           (numpar, 1)), axis=1)-Tpar[rws]
        Rth = 0.1
        maxtry = 20
        cnt = 0
        while(np.amax(np.absolute(R)) > Rth and cnt < maxtry):
            xmax_rp = np.tile(xmax, (numpar, 1))
            J = -ptpar_rp*Xg_sub/xmax_rp/xmax_rp
            Jinv = np.linalg.inv(J)
            R = np.sum(ptpar_rp*Xg_sub/xmax_rp, axis=1)-Tpar[rws]
            xmax = xmax-np.matmul(Jinv, R)
            cnt += 1
        for ii in range(numpar):
            qans[par_indx[ii]] = int(round(xmax[ii]))

    # convert values to strings
    for jj in range(n):
        qval[jj] = str(qval[jj])

    # assemble the header
    teks.insert(0, 'TEK')
    qval.insert(0, 'value')
    qans.insert(0, 'answer')
    qtype.insert(0, 'type')
    hdr = [teks, qval, qans, qtype]

    return hdr


def ndatis_fname(tchr, crs, per, date):
    '''ndatis_fname constructs the Ndatis file name
    Parameters:
    tchr, name of teacher (<first>_<middle initial>_<last>)
    crs, the course (algebra 1,algebra 2, or precal)
    per, the period 
    date, date of assessment (<two-digit month><two-digit day><four-digit year>)    

    Returns:
    fname, the name of the Ndatis file (extension .csv)
    '''

    crsabbr = {'algebra1': 'alg1', 'algebra2': 'alg2', 'precal': 'pcal'}
    tmp = tchr.split('_')
    tchtag = tmp[0][0]+tmp[1][0]+tmp[2][0]
    pnum = per.strip('period')
    fname = 'f_asmt_'+date+'_'+tchtag+crsabbr.get(crs)+'per'+pnum+'.csv'

    return fname


def build_ndatisfile(X, names, hdr):
    '''build_ndatisfile build a file for NDatis

    Parameters:
    X, student responses (# student x # questions)
    names, student names (# student x 1)
    hdr, file header (4 x # questions)

    Returns:
    data for Ndatis file
    '''

    [X[li].insert(0, names[li]) for li in range(len(names))]
    hdr.extend(X)

    return hdr


to = time.time()

# START HERE
''' ==================== Find Files ==================== '''

fnames = glob.glob(datadir+'/*.xlsx')
numfiles = len(fnames)


''' ==================== Read Exported Data ==================== '''

alldata = [None]*numfiles
for ii in range(numfiles):
    book = openpyxl.load_workbook(fnames[ii])
    sheet = book.active  # sheet 1 (default)
    alldata[ii] = extrdata(sheet)


''' ==================== Reorganize Data ==================== '''

# preallocate
X = []
snames = []
score = []
perc = []

# the TEKS are the same for all files
# => take TEKS from only first file
teks = alldata[0].get('teks')

# accumulate data from multiple files
for ii in range(numfiles):
    X = X+alldata[ii].get('responses')
    snames = snames+alldata[ii].get('names')
    score = score+alldata[ii].get('raw scores')
    perc = perc+alldata[ii].get('percent scores')


''' ==================== Construct Data for Import ==================== '''

fnout = ndatis_fname(teacher, course, period, asmtdate)
filehdr = buildhead(X, teks, score, perc)

if(deid == True):  # de-identify
    studyids = ['']*len(snames)
    allnames, allids = readroster(rostdir+'/'+rostfile)
    for ii in range(len(snames)):
        eindx = ele_in_list(snames[ii], allnames)
        if(eindx is None):
            print('roster is missing names')
            exit()
        else:
            studyids[ii] = allids[eindx]
    impdata = build_ndatisfile(X, studyids, filehdr)
else:  # don't de-identify
    impdata = build_ndatisfile(X, snames, filehdr)


''' ==================== Save Data ==================== '''

tf = time.time()
print('file(s) converted in', tf-to, 'seconds')

csvw.write_csv(impdata, datadir+'/'+fnout)
